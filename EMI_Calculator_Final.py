from Tkinter import *
import openpyxl
from openpyxl import *
from openpyxl.styles import *
from openpyxl.cell import *
from openpyxl.chart import *
from openpyxl.workbook import *
from openpyxl.chart.series import DataPoint

class emi_calculator():
    def __init__(self):
        self.root = Tk()
        self.root.title("EMI Calculator")

        Label(self.root, text="Principal (in Rs): ",font = "TrebuchetMS 10 bold").grid(row=0, column=0, sticky=W)
        self.P = DoubleVar()
        Entry(self.root, textvariable=self.P).grid(row=0, column=1, sticky=W)

        Label(self.root, text="Rate of Interest (in %): ",font = "TrebuchetMS 10 bold").grid(row=2, column=0, sticky=W)
        self.R = DoubleVar()
        Entry(self.root, textvariable=self.R).grid(row=2, column=1, sticky=W)

        Label(self.root, text="EMI Tenure (Year(s)/Month(s)): ",font = "TrebuchetMS 10 bold").grid(row=4, column=0, sticky=W)
        self.y = IntVar()
        Entry(self.root, textvariable=self.y, width=10).grid(row=4, column=1, sticky=W)
        self.m = IntVar()
        Entry(self.root, textvariable=self.m, width=10).grid(row=4, column=1, sticky=E)

        Label(self.root, text="Select Ouput File",font = "TrebuchetMS 10 bold").grid(row=8, column=0, sticky=W)
        self.v1 = IntVar()
        Radiobutton(self.root, text="Excel", indicatoron = 1, width = 0, padx = 20,  variable=self.v1, command=self.v1.get(), value="1").grid(row=10, column=0, sticky=W)
        self.v2 = IntVar()
        Radiobutton(self.root, text="PDF", indicatoron = 1, width = 0, padx = 20,  variable=self.v2, command=self.v2.get(), value="1").grid(row=12, column=0, sticky=W)
        self.v3 = IntVar()
        Radiobutton(self.root, text="Just Display in Command Prompt", indicatoron = 1, width = 0, padx = 20,  variable=self.v3, command=self.v3.get(), value="1").grid(row=14, column=0, sticky=W)

        Button(self.root, text='Calculate', width=20, command=self.Calculate).grid(row=18, column=0, sticky=W)
        Button(self.root, text='Quit', width=20, command=self.root.destroy).grid(row=20, column=0, sticky=W)
        self.root.mainloop()

    def Calculate(self):
        self.P = self.P.get()
        self.R = self.R.get()
        self.y = self.y.get()
        self.m = self.m.get()
        self.n = ((self.y)*12)+(self.m)
        self.excel = self.v1.get()
        self.pdf = self.v2.get()
        self.display = self.v3.get()
        if (self.excel==1):
            self.excel_func()
        elif (self.pdf==1):
            self.pdf_func()
        elif (self.display==1):
            self.display_func()
        else:
            print ("Need to Select One Option.")

    def excel_func(self):
        P = self.P
        R = self.R
        n = self.n
        
        #year variable is considered  for advanced calculation
        year=12

        #roi is taken for advanced calculation
        roi=R/100

        #temp_1 & temp_2 are considered for calculations
        temp_1=(R/12/100)
        temp_2=((1+temp_1)**n)

        #Processing Fee
        Proc_temp=(0.02*P)
        Proc=round(Proc_temp, 2)

        #EMI Amount
        EMI=((P*temp_1*temp_2)/(temp_2-1))

        #Following values are complementry functions which can be neglected but are important for advance calculations
        total_amount=EMI*n
        interest=total_amount-P

        #Advance Calculation begins here

        #This calculation is done only for 1st installment
        Int1=(P*roi)/year
        Pri1=EMI-Int1
        Pen1=P-Pri1

        #Rounding of the values of 1st installment to 2 decimal places
        Inter= round(Int1, 2)
        Princ=round(Pri1, 2)
        Pendi=round(Pen1, 2)

        matrix_1 = [Inter, Princ, Pendi]
        #print("\nInstallment Number: 1")
        #print (matrix_1)

        #this temporary variable is considered  since interest requires pending of previous loop
        temp_7=Pen1

        #to initialise excel sheet
        wb = Workbook()
        ws = wb.get_sheet_by_name(name = 'Sheet')
        ws.column_dimensions["A"].width = 19
        ws.column_dimensions["B"].width = 19
        ws.column_dimensions["C"].width = 19
        ws.column_dimensions["D"].width = 19
        ws.column_dimensions["E"].width = 19

        #To set various styles
        font1 = Font(name='Trebuchet MS', size=10, bold=True, italic=False, vertAlign=None, underline='none', strike=False, color='FF000000')
        font2 = Font(name='Trebuchet MS', size=10, bold=False, italic=False, vertAlign=None, underline='none', strike=False, color='FF000000')
        fill1 = PatternFill(fill_type='solid', start_color='FF0000', end_color='FF0000')
        border1 = Border(left=Side(border_style='thin', color='00000000'), right=Side(border_style='thin', color='00000000'), top=Side(border_style='thin', color='00000000'),bottom=Side(border_style='thin', color='00000000'), diagonal=Side(border_style=None,
         color='FF000000'), diagonal_direction=0, outline=Side(border_style=None, color='FF000000'), vertical=Side(border_style=None, color='FF000000'), horizontal=Side(border_style=None, color='FF000000'))
        alignment1 =Alignment(horizontal='center', vertical='center', text_rotation=0, wrap_text=True, shrink_to_fit=False, indent=0)
        alignment2 =Alignment(horizontal='left', vertical='center', text_rotation=0, wrap_text=True, shrink_to_fit=False, indent=0)
        alignment3 =Alignment(horizontal='right', vertical='center', text_rotation=0, wrap_text=True, shrink_to_fit=False, indent=0)
        protection1 = Protection(locked=True, hidden=False)
        currency_format="#,##0.00;[Red]-#,##0.00"

        #this is to write header file in the top of excel
        ws.cell(row=1, column=1).value = "Principal Amount(Rs):"
        ws.cell(row=2, column=1).value = "Rate of Interest (%):"
        ws.cell(row=3, column=1).value = "Tenure (Months):"
        ws.cell(row=5, column=1).value = "Processing Fee"
        ws.cell(row=7, column=1).value = "Total Amount Paid:(Including Principle Amount & Interest)"
        ws.cell(row=8, column=1).value = "Total Interest paid: "

        ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=3)
        ws.merge_cells(start_row=2,start_column=1,end_row=2,end_column=3)
        ws.merge_cells(start_row=3,start_column=1,end_row=3,end_column=3)
        ws.merge_cells(start_row=4,start_column=1,end_row=4,end_column=3)
        ws.merge_cells(start_row=5,start_column=1,end_row=5,end_column=3)
        ws.merge_cells(start_row=6,start_column=1,end_row=6,end_column=3)
        ws.merge_cells(start_row=7,start_column=1,end_row=7,end_column=3)
        ws.merge_cells(start_row=8,start_column=1,end_row=8,end_column=3)
        
        ws.cell(row=1, column=4).value = ("Rs. {0:.2f}".format(P))
        ws.cell(row=2, column=4).value = ("{0:.2f} %".format(R))
        ws.cell(row=3, column=4).value = ("{0:d} Months".format(n))
        ws.cell(row=5, column=4).value = ("Rs. {0:.2f}".format(Proc))
        ws.cell(row=7, column=4).value = ("Rs. {0:.2f}".format(total_amount))
        ws.cell(row=8, column=4).value = ("Rs. {0:.2f}".format(interest))

        #this is to write header content of table
        ws.cell(row=15, column=1).value = "Installment Number"
        ws.cell(row=15, column=2).value = "Principal"
        ws.cell(row=15, column=3).value = "Interest"
        ws.cell(row=15, column=4).value = "EMI Amount"
        ws.cell(row=15, column=5).value = "Pending Amount"

        #this is to write details of 1st installment since formula is bit different
        inst_no=1
        xyz = 16
        ws.cell(row=xyz, column=1).value = inst_no
        ws.cell(row=xyz, column=2).value = Princ
        ws.cell(row=xyz, column=3).value = Inter
        ws.cell(row=xyz, column=4).value = (Princ+Inter)
        ws.cell(row=xyz, column=5).value = Pendi
        xyz+=1

        #xyz is considered for looping purpose.
        inst_no = 2
        while inst_no<=n:
            Inte=(temp_7*roi)/year
            Prin=EMI-Inte
            Pend=temp_7-Prin

            #All the float values are rounded of to 2 decimal places
            Intere= round(Inte, 2)
            Princi=round(Prin, 2)
            Pendin=round(Pend, 2)

            #Data (which was rounded of to 2 Decimal places) is written in specific cells.
            ws.cell(row=xyz, column=1).value = inst_no
            ws.cell(row=xyz, column=2).value = Princi
            ws.cell(row=xyz, column=3).value = Intere
            ws.cell(row=xyz, column=4).value = (Princi+Intere)
            ws.cell(row=xyz, column=5).value = Pendin

            temp_7=Pend
            xyz+=1
            inst_no+=1
            
        #For final totaling of Principal Amount and Interest
        ws.cell(row=xyz, column=1).value = "Total"
        ws.cell(row=xyz, column=2).value = P
        ws.cell(row=xyz, column=3).value = round(interest, 2)
        ws.cell(row=xyz, column=4).value = "---"
        ws.cell(row=xyz, column=5).value = Pendin
        row_temp=1
        while (row_temp<=15):
            column_temp=1
            while (column_temp<=5):
                ws.cell(row=row_temp, column=column_temp).font = font1
                ws.cell(row=row_temp, column=column_temp).alignment = alignment2
                column_temp+=1
            row_temp+=1

        row_temp=15
        while (row_temp<=xyz):
            column_temp=1
            while (column_temp<=5):
                ws.cell(row=row_temp, column=column_temp).alignment = alignment1
                ws.cell(row=row_temp, column=column_temp).border=border1
                column_temp+=1
            row_temp+=1

        row_temp=16
        while (row_temp<=xyz):
            column_temp=1
            while (column_temp<=5):
                ws.cell(row=row_temp, column=column_temp).font = font2
                column_temp+=1
            row_temp+=1

        column_temp=1
        while (column_temp<=5):
            ws.cell(row=(xyz), column=column_temp).font = font1
            column_temp+=1

        row_temp=1
        while (row_temp<=8):
            ws.cell(row=row_temp, column=4).alignment = alignment3
            row_temp+=1

        row_temp=16
        while (row_temp<=xyz):
            column_temp=2
            while (column_temp<=5):
                ws.cell(row=row_temp, column=column_temp).number_format = currency_format
                column_temp+=1
            row_temp+=1
        
        #ws.cell(row=1, column=2).number_format = currency_format
        #ws.cell(row=5, column=2).number_format = currency_format
        #ws.cell(row=7, column=2).number_format = currency_format
        #ws.cell(row=8, column=2).number_format = currency_format
        
        data = Reference(ws, min_col=2, min_row=15, max_col=3, max_row=(xyz-1))
        titles = Reference(ws, min_col=1, max_col=1, min_row=16, max_row=(xyz-1))
        chart = BarChart()
        chart.title = "Detailed Monthly Report"
        chart.x_axis.title = 'Installment Number'
        chart.y_axis.title = 'Monthly EMI'
        chart.add_data(data=data, titles_from_data=True)
        chart.set_categories(titles)
        chart.height = 15.88
        chart.width = 31.3
        chart.grouping = "stacked"
        chart.overlap = 100
        ws.add_chart(chart, "G17")

        pie = PieChart()
        labels = Reference(ws, min_col=2, max_col=3, min_row=15, max_row=15)
        data = Reference(ws, min_col=2, max_col=3, min_row=40, max_row=40)
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        pie.height = 7.41
        pie.width = 11.85
        pie.title = "Total Principal to Interest Ratio"

        # Cut the first slice out of the pie
        slice = DataPoint(idx=0, explosion=20)
        pie.series[0].data_points = [slice]

        ws.add_chart(pie, "G1")
    
        wb.save("EMI_{0:.2f}_{1:d}_{2:.2f}.xlsx".format(P, n, R))
        print ("Excel file is generated.")
        self.root.destroy()

    def pdf_func(self):
        print "Under Progress"
        self.root.destroy()

    def display_func(self):
        P = self.P
        R = self.R
        n = self.n
        
        #year variable is considered  for advanced calculation
        year=12

        #roi is taken for advanced calculation
        roi=R/100

        #temp_1 & temp_2 are considered for calculations
        temp_1=(R/12/100)
        temp_2=((1+temp_1)**n)

        #Processing Fee
        Proc_temp=(0.02*P)
        Proc=round(Proc_temp, 2)

        #EMI Amount
        EMI=((P*temp_1*temp_2)/(temp_2-1))

        #printing values of principal, rate of interest, tenure & final calculated EMI
        print("Principle Amount: Rs. %d" % P)
        print("Rate of Interest: {0:.2f}%".format(R))
        print("Tenure: %d Months" % n)
        print("\nTotal Calulated EMI is: Rs. %.2f" % EMI)

        #Following values are complementry functions which can be neglected but are important for advance calculations
        total_amount=EMI*n
        print("\nTotal Amount Paid: Rs. %.2f (Including Principle Amount & Interest)" % total_amount)
        interest=total_amount-P
        print("Interest paid: Rs. %.2f" % interest)
        self.root.destroy()

obj=emi_calculator()
